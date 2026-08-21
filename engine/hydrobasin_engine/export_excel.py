from __future__ import annotations

from pathlib import Path
from typing import Any
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def export_ideam_and_hydrology_excel(
    output_excel_path: Path,
    summary: dict[str, Any],
) -> Path | None:
    """Genera el libro de cálculo integral 'resultados_hidrologicos.xlsx' con el registro completo de:
      1. Unidades Homogéneas CN (CORINE 2018 + Geología SGC + CN II).
      2. Estaciones Meteorológicas IDEAM.
      3. Polígonos de Thiessen y Ponderación Territorial.
      4. Curvas IDF y Lluvias de Diseño.
      5. Caudales Máximos de Diseño por Periodo de Retorno.
      6. Series Temporales de Hidrogramas Q(t).
    """
    output_excel_path = Path(output_excel_path).resolve()
    output_excel_path.parent.mkdir(parents=True, exist_ok=True)

    stations = summary.get("ideam_stations") or []
    thiessen = summary.get("thiessen_weights") or []
    peaks = summary.get("peak_discharges") or []
    hydro = summary.get("hydrologic_modeling", {}).get("hydrographs") or {}
    cn_data = summary.get("curve_number") or {}
    cn_units = cn_data.get("units") or summary.get("cn_units") or []

    # Escribir en un archivo temporal o directo
    with pd.ExcelWriter(str(output_excel_path), engine="openpyxl") as writer:
        # 1. Hoja: Unidades Homogéneas de Número de Curva SCS (CN II)
        if cn_units:
            u_rows = []
            for u in cn_units:
                u_rows.append({
                    "Cobertura CORINE 2018": u.get("cobertura") or "N/D",
                    "Uso SCS": u.get("uso_scs") or "No clasificado",
                    "Condición Hidrológica": u.get("condicion") or "N/D",
                    "Símbolo Geológico SGC": u.get("simbolo_uc") or "N/D",
                    "Litología / Formación SGC": u.get("litologia") or "N/D",
                    "Grupo Hidrológico HSG": u.get("grupo_suelo") or "No clasificado",
                    "Número de Curva (CN II)": u.get("cn") if u.get("cn") is not None else "N/D",
                    "Área (km²)": u.get("area_km2", 0.0),
                    "Porcentaje Cuenca (%)": u.get("porcentaje_cuenca", 0.0),
                    "CN × Área (km²)": u.get("nc_ai", 0.0),
                })
            df_cn = pd.DataFrame(u_rows)
        else:
            df_cn = pd.DataFrame([{"Mensaje": "Sin unidades CN calculadas"}])
        df_cn.to_excel(writer, sheet_name="Unidades_Homogeneas_CN", index=False)

        # 2. Hoja: Estaciones IDEAM
        st_data = []
        for s in stations:
            st_data.append({
                "Código Estación": s.get("codigo") or "N/D",
                "Nombre": s.get("nombre") or "N/D",
                "Categoría": s.get("categoria") or "N/D",
                "Tecnología": s.get("tecnologia") or "Convencional",
                "Estado Operativo": s.get("estado") or "Activa",
                "Departamento": s.get("departamento") or "N/D",
                "Municipio": s.get("municipio") or "N/D",
                "Altitud (msnm)": s.get("altitud") if s.get("altitud") is not None else "N/D",
                "Latitud (°)": s.get("latitud") if s.get("latitud") is not None else "N/D",
                "Longitud (°)": s.get("longitud") if s.get("longitud") is not None else "N/D",
                "Distancia (km)": s.get("distancia_km") if s.get("distancia_km") is not None else "N/D",
                "Entidad": s.get("entidad") or "IDEAM",
            })
        df_st = pd.DataFrame(st_data) if st_data else pd.DataFrame([{"Mensaje": "Sin estaciones registradas"}])
        df_st.to_excel(writer, sheet_name="Estaciones_IDEAM", index=False)

        # 3. Hoja: Polígonos de Thiessen
        th_data = []
        for th in thiessen:
            th_data.append({
                "Código Estación": th.get("codigo") or "N/D",
                "Nombre": th.get("nombre") or "N/D",
                "Área de Influencia (km²)": th.get("area_km2", 0.0),
                "Porcentaje de Cuenca (%)": th.get("porcentaje", 0.0),
            })
        df_th = pd.DataFrame(th_data) if th_data else pd.DataFrame([{"Mensaje": "Ponderación uniforme"}])
        df_th.to_excel(writer, sheet_name="Poligonos_Thiessen", index=False)

        # 4. Hoja: Curvas IDF y Lluvias de Diseño
        idf_data = []
        for q in peaks:
            idf_data.append({
                "Periodo de Retorno Tr (años)": q.get("tr_anos"),
                "Intensidad de Diseño I (mm/h)": q.get("intensidad_mm_h"),
                "Precipitación Total P (mm)": q.get("precipitacion_total_mm"),
                "Precipitación Efectiva Pe (mm)": q.get("precipitacion_efectiva_mm"),
            })
        df_idf = pd.DataFrame(idf_data) if idf_data else pd.DataFrame([{"Mensaje": "Sin datos IDF"}])
        df_idf.to_excel(writer, sheet_name="Curvas_IDF_Lluvias", index=False)

        # 5. Hoja: Caudales Máximos de Diseño
        q_data = []
        for q in peaks:
            q_data.append({
                "Periodo de Retorno Tr (años)": q.get("tr_anos"),
                "Intensidad I (mm/h)": q.get("intensidad_mm_h"),
                "P Total (mm)": q.get("precipitacion_total_mm"),
                "P Efectiva (mm)": q.get("precipitacion_efectiva_mm"),
                "Caudal Método Racional (m³/s)": q.get("caudal_racional_m3_s"),
                "Caudal Método SCS (m³/s)": q.get("caudal_scs_m3_s"),
                "Caudal de Diseño Adoptado (m³/s)": q.get("caudal_diseno_m3_s"),
            })
        df_q = pd.DataFrame(q_data) if q_data else pd.DataFrame([{"Mensaje": "Sin caudales"}])
        df_q.to_excel(writer, sheet_name="Caudales_Diseno", index=False)

        # 6. Hoja: Series Temporales de Hidrogramas Q(t)
        if hydro:
            first_key = next(iter(hydro.keys()))
            times_h = hydro[first_key].get("time_hours", [])
            times_min = hydro[first_key].get("time_minutes", [])
            ts_dict: dict[str, list[Any]] = {
                "Tiempo (horas)": times_h,
                "Tiempo (minutos)": times_min,
            }
            for k, v in hydro.items():
                col_name = f"Q {k.replace('_', ' = ')} años (m³/s)"
                ts_dict[col_name] = v.get("flow_m3_s", [])
            df_ts = pd.DataFrame(ts_dict)
            df_ts.to_excel(writer, sheet_name="Series_Hidrogramas_Q_t", index=False)

    # Estilizar el archivo Excel con OpenPyXL
    wb = openpyxl.load_workbook(str(output_excel_path))
    header_fill = PatternFill(start_color="1F9D8F", end_color="1F9D8F", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=10, color="0F172A")
    alt_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )

    for sheet in wb.worksheets:
        sheet.views.sheetView[0].showGridLines = True
        for col_idx, cell in enumerate(sheet[1], start=1):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row_idx in range(2, sheet.max_row + 1):
            is_alt = row_idx % 2 == 1
            for col_idx in range(1, sheet.max_column + 1):
                c = sheet.cell(row=row_idx, column=col_idx)
                c.font = data_font
                c.border = thin_border
                if is_alt:
                    c.fill = alt_fill
                if isinstance(c.value, (int, float)):
                    c.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    c.alignment = Alignment(horizontal="left", vertical="center")

        # Ajuste automático de anchos de columna
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or "")
                if len(val_str) > max_len:
                    max_len = len(val_str)
            sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

    wb.save(str(output_excel_path))

    # Guardar copia compatible 'estaciones_ideam.xlsx' y 'resultados_hidrologicos.xlsx'
    alt_name = "resultados_hidrologicos.xlsx" if output_excel_path.name == "estaciones_ideam.xlsx" else "estaciones_ideam.xlsx"
    alt_path = output_excel_path.parent / alt_name
    try:
        wb.save(str(alt_path))
    except Exception:
        pass

    return output_excel_path
